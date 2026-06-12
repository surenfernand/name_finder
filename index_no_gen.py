from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "Index Numbers"

# Header
ws["A1"] = "index_number"

start_number = 96019830 
count = 100000  # change as needed

for i in range(count):
    ws.cell(row=i + 2, column=1, value=start_number + i)

wb.save("indexes.xlsx")

print(f"Created indexes.xlsx with {count} numbers")